"""
Table Convertor Module
exports mysql Tables to different types of files
"""

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from idLogTable import idLogTable

def saveAsTXT(table, path):
    try:
        outputFile = open(path, 'w+')
        outputFile.write(str(table))
        outputFile.close()
    except:
        print 'Output file could not be opened'
    return

def saveAsXLSX(table, path):
    wb = Workbook()
    entrySheet = wb.worksheets[0]
    entrySheet.title = 'Entries'
    entrySheet.cell('A1').value = 'Alp'
    
    for col_idx in xrange(1, table.getHeadersLength()+1):
        col = get_column_letter(col_idx)
        entrySheet.cell('%s%s'%(col, 1)).value = str(table.getHeaders()[col_idx-1])
    
    for row in xrange(2,len(table)+2):
        rowList = table.getRow(row-2)
        for col_idx in xrange(1,len(rowList)+1):
            col = get_column_letter(col_idx)
            entrySheet.cell('%s%s'%(col, row)).value = str(rowList[col_idx-1])
    
    wb.save(filename = path)


def saveAsPDF(table, path):
    return

def saveAsXML(table, path):
    return

def saveAsHTML(table, path):
    return

if __name__ == "__main__":
    print 'hello'
    dummyTable = idLogTable(['id', 'username', 'category', 'blabla'])
    dummyTable.setHeaders(['id', 'username', 'category', 'content'])
    dummyTable.addRow([1, 'alpsayin', 'farnell', 'raspberry pi siparisleri verildi'])
    dummyTable.addRow([2, 'alpsayin', 'farnell', 'raspberry pi siparisleri elimize ulasti'])
    dummyTable.addRow([3, 'umutgultepe', 'g-man', 'some entry'], 6)
#    dummyTable.printTable()
    saveAsTXT(dummyTable, 'alp.txt')
    saveAsXLSX(dummyTable, 'alp.xlsx')
    print 'bye bye'